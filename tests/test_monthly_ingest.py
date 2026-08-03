"""Tests for ``monthly_report_ingest`` and ``monthly_report_xlsx``.

Hermetic: the fake Supabase stands in for the database and a real POS fixture
supplies the content, so the round trip parse -> store -> read back is exercised
without a network.
"""

import os
import sys
import tempfile
import unittest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from tests.fake_supabase import FakeSupabase  # noqa: E402
from monthly_report_ingest import (  # noqa: E402
    ITEMWISE_TABLE,
    MONTHLY_REPORT_TABLE,
    PAYOUTS_TABLE,
    STAFF_SALES_TABLE,
    _sale_date,
    build_rows,
    ingest_monthly_email,
    resolve_outlet,
    store_monthly_report,
)
from monthly_report_parser import parse_monthly_report  # noqa: E402
from sales_email_fetcher import detect_email_type  # noqa: E402
from sales_parser import read_shift_close_file  # noqa: E402

REAL_POS_FILE = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "fixtures", "sales",
    "S-Damansara 25May2026.TXT",
)


def monthly_content():
    """Real POS body under a monthly header — same generator, same grammar."""
    return "MONTHLY REPORT-Damansara ON Jul 2026\n" + read_shift_close_file(REAL_POS_FILE)


class SubjectDetection(unittest.TestCase):

    def test_monthly_subjects_are_recognised(self):
        self.assertEqual(detect_email_type("MONTHLY REPORT-Damansara ON Jul 2026"),
                         ("M", "DAMANSARA"))
        self.assertEqual(detect_email_type("M-Damansara ON Jul 2026"), ("M", "DAMANSARA"))

    def test_daily_and_shift_subjects_are_unchanged(self):
        self.assertEqual(detect_email_type("S-KLANG SHIFTCLOSE 2386"), ("S", "S-KLANG"))
        self.assertEqual(detect_email_type("D-SEK20 ON 26/May/2026"), ("D", "D-SEK20"))
        self.assertEqual(detect_email_type("D-ST KHU ON 26/May/2026"), ("D", "D-ST KHU"))

    def test_unrelated_mail(self):
        self.assertIsNone(detect_email_type("Your invoice is ready"))
        self.assertIsNone(detect_email_type(None))


class OutletResolution(unittest.TestCase):

    def test_header_code_resolves_to_canonical(self):
        self.assertEqual(resolve_outlet("DAMANSARA", None), "D.U")
        self.assertEqual(resolve_outlet("KLANG", None), "Klang B.Emas")

    def test_header_wins_over_subject(self):
        self.assertEqual(resolve_outlet("KLANG", "DAMANSARA"), "Klang B.Emas")

    def test_unmappable_code_is_kept_not_dropped(self):
        self.assertEqual(resolve_outlet("BRAND NEW SHOP", None), "BRAND NEW SHOP")

    def test_nothing_at_all(self):
        self.assertEqual(resolve_outlet(None, None), "UNKNOWN")


class DailyDateCells(unittest.TestCase):

    def test_shapes_within_the_period(self):
        for cell in ("01", "1/7", "01/07/2026"):
            self.assertEqual(_sale_date(cell, "2026-07"), "2026-07-01", cell)

    def test_a_row_from_another_month_is_not_re_dated(self):
        # Silently filing August's sales under July would be worse than a NULL.
        self.assertIsNone(_sale_date("15/08/2026", "2026-07"))

    def test_junk(self):
        for cell in ("", None, "abc", "99"):
            self.assertIsNone(_sale_date(cell, "2026-07"), cell)


class BuildRows(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        cls.parsed = parse_monthly_report(monthly_content())
        cls.rows = build_rows(cls.parsed, "D.U", "2026-07")

    def test_every_table_is_keyed_by_outlet_and_period(self):
        for table, rows in self.rows.items():
            for row in rows:
                self.assertEqual(row["outlet"], "D.U", table)
                self.assertEqual(row["period"], "2026-07", table)

    def test_payout_blocks_are_labelled(self):
        # The S- shift fixture uses shift closing labels, so the monthly
        # closing-total anchors claim nothing from it — no money-block rows.
        self.assertEqual(self.rows[PAYOUTS_TABLE], [])

    def test_itemwise_keeps_the_verbatim_name(self):
        names = [r["item_name"] for r in self.rows[ITEMWISE_TABLE]]
        self.assertIn(" Nasi Goreng Daging", names)
        self.assertIn("Barli  Panas", names)

    def test_line_numbers_are_unique_within_a_block(self):
        for block in ("purchase", "pinjam"):
            numbers = [r["line_no"] for r in self.rows[PAYOUTS_TABLE] if r["block"] == block]
            self.assertEqual(len(numbers), len(set(numbers)), block)

    def test_repeated_staff_rows_are_summed_not_collided(self):
        parsed = {"staff_sales": [{"staff_name": "HARI", "amount": 100.0},
                                  {"staff_name": "HARI", "amount": 50.0}]}
        rows = build_rows(parsed, "D.U", "2026-07")[STAFF_SALES_TABLE]
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["amount"], 150.0)


class StoreAndReIngest(unittest.TestCase):

    def setUp(self):
        self.db = FakeSupabase()
        self.parsed = parse_monthly_report(monthly_content())

    def test_stores_header_and_detail(self):
        result = store_monthly_report(self.db, self.parsed, outlet="D.U", period="2026-07")
        self.assertIsNotNone(result["monthly_report_id"])
        self.assertEqual(len(self.db.rows(ITEMWISE_TABLE)), 143)
        header = self.db.rows(MONTHLY_REPORT_TABLE)[0]
        self.assertEqual(header["outlet"], "D.U")
        self.assertAlmostEqual(header["total_payouts"], 1024.0, places=2)

    def test_re_ingesting_replaces_rather_than_duplicates(self):
        store_monthly_report(self.db, self.parsed, outlet="D.U", period="2026-07")
        store_monthly_report(self.db, self.parsed, outlet="D.U", period="2026-07")
        self.assertEqual(len(self.db.rows(ITEMWISE_TABLE)), 143)
        self.assertEqual(len(self.db.rows(MONTHLY_REPORT_TABLE)), 1)

    def test_another_month_does_not_clear_this_one(self):
        store_monthly_report(self.db, self.parsed, outlet="D.U", period="2026-07")
        store_monthly_report(self.db, self.parsed, outlet="D.U", period="2026-08")
        periods = {r["period"] for r in self.db.rows(ITEMWISE_TABLE)}
        self.assertEqual(periods, {"2026-07", "2026-08"})

    def test_detail_rows_link_to_the_header(self):
        result = store_monthly_report(self.db, self.parsed, outlet="D.U", period="2026-07")
        report_id = result["monthly_report_id"]
        self.assertTrue(all(r["monthly_report_id"] == report_id
                            for r in self.db.rows(ITEMWISE_TABLE)))


class IngestEmail(unittest.TestCase):

    def _email(self, content, **kwargs):
        return dict({"content": content, "filename": "M.TXT", "message_id": "<1>",
                     "outlet_code": "DAMANSARA",
                     "subject": "MONTHLY REPORT-Damansara ON Jul 2026"}, **kwargs)

    def test_happy_path(self):
        db = FakeSupabase()
        result = ingest_monthly_email(db, self._email(monthly_content()))
        self.assertEqual(result["status"], "ok")
        self.assertEqual(result["outlet"], "D.U")
        self.assertEqual(result["period"], "2026-07")
        self.assertEqual(result["counts"][ITEMWISE_TABLE], 143)

    def test_empty_attachment_is_an_error_not_a_silent_success(self):
        result = ingest_monthly_email(FakeSupabase(), self._email(""))
        self.assertEqual(result["status"], "error")

    def test_content_without_a_monthly_header_is_skipped(self):
        result = ingest_monthly_email(
            FakeSupabase(), self._email("S-Damansara ON 25/May/2026\nTOTAL SALES : 10.00")
        )
        self.assertEqual(result["status"], "skipped")

    def test_a_store_failure_is_reported_not_raised(self):
        class Broken:
            def table(self, _name):
                raise RuntimeError("upstream down")

        result = ingest_monthly_email(Broken(), self._email(monthly_content()))
        self.assertEqual(result["status"], "error")
        self.assertIn("upstream down", result["reason"])


class Workbook(unittest.TestCase):

    def setUp(self):
        from monthly_close import build_cash_reconciliation, build_pl
        from menu_hygiene import build_menu_hygiene
        from monthly_wastage import build_wastage

        parsed = parse_monthly_report(monthly_content())
        totals = {"total_sales": 148586.80, "total_bank_sales": 70000.0,
                  "total_fp_sales": 8948.30, "total_payouts": 45121.10}
        self.pl = build_pl(totals, [], [], [{"description": "BALAJI", "amount": 1000.0}],
                           set(), None)
        self.cash = build_cash_reconciliation(totals)
        self.hygiene = build_menu_hygiene(parsed["itemwise"])
        self.wastage = build_wastage(parsed["itemwise"], [])
        self.dir = tempfile.mkdtemp()

    def _write(self, path):
        from monthly_report_xlsx import write_monthly_sheets

        return write_monthly_sheets(path, outlet="D.U", period="2026-07", pl=self.pl,
                                    cash=self.cash, hygiene=self.hygiene,
                                    wastage=self.wastage)

    def test_filename_matches_the_report_convention(self):
        from monthly_report_xlsx import report_filename

        self.assertEqual(report_filename("Klang B.Emas", "2026-07"),
                         "Khulafa_KLANG_B_EMAS_JULY_2026_Wastage_Report.xlsx")

    def test_creates_the_workbook_with_the_four_sheets(self):
        from openpyxl import load_workbook

        path = self._write(os.path.join(self.dir, "new.xlsx"))
        sheets = load_workbook(path).sheetnames
        for name in ("P&L", "Cash Reconciliation", "Menu Hygiene", "Wastage Variance"):
            self.assertIn(name, sheets)

    def test_existing_sheets_are_preserved(self):
        from openpyxl import Workbook as WB, load_workbook

        path = os.path.join(self.dir, "existing.xlsx")
        wb = WB()
        wb.active.title = "Cost Calculator"
        wb["Cost Calculator"]["A1"] = "owner edits here"
        wb.save(path)

        self._write(path)
        reopened = load_workbook(path)
        self.assertIn("Cost Calculator", reopened.sheetnames)
        self.assertEqual(reopened["Cost Calculator"]["A1"].value, "owner edits here")

    def test_rewriting_does_not_stack_two_runs_of_rows(self):
        from openpyxl import load_workbook

        path = os.path.join(self.dir, "twice.xlsx")
        self._write(path)
        first = load_workbook(path)["P&L"].max_row
        self._write(path)
        self.assertEqual(load_workbook(path)["P&L"].max_row, first)

    def test_contribution_label_when_config_is_missing(self):
        from openpyxl import load_workbook

        path = self._write(os.path.join(self.dir, "contrib.xlsx"))
        ws = load_workbook(path)["P&L"]
        labels = " ".join(str(row[0]) for row in ws.iter_rows(values_only=True) if row[0])
        self.assertIn("CONTRIBUTION (pre-rent/utilities/central payroll)", labels)
        self.assertNotIn("NET PROFIT", labels)

    def test_cash_sheet_carries_the_expected_bank_in(self):
        from openpyxl import load_workbook

        path = self._write(os.path.join(self.dir, "cash.xlsx"))
        ws = load_workbook(path)["Cash Reconciliation"]
        values = {row[0]: row[1] for row in ws.iter_rows(values_only=True) if row[0]}
        self.assertAlmostEqual(
            values["EXPECTED BANK-IN (match this against the bank statement)"],
            24517.40, places=2,
        )


if __name__ == "__main__":
    unittest.main()


class IntegrityGuardBlocksStorage(unittest.TestCase):
    """A month that does not reconcile is never stored, and carries an alert."""

    def _content(self):
        path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "fixtures",
                            "monthly_synthetic", "M-Damansara Jul2026-SYNTHETIC.TXT")
        return read_shift_close_file(path)

    def _email(self, content):
        return {"content": content, "filename": "M.TXT", "message_id": "<1>",
                "outlet_code": "DAMANSARA",
                "subject": "MONTHLY REPORT-Damansara ON Jul 2026"}

    def test_a_reconciling_month_stores(self):
        db = FakeSupabase()
        result = ingest_monthly_email(db, self._email(self._content()))
        self.assertEqual(result["status"], "ok")
        self.assertGreater(len(db.rows(PAYOUTS_TABLE)), 0)

    def test_a_broken_month_is_rejected_and_nothing_is_written(self):
        # Corrupt the supplier block's closing total so its rows no longer add up.
        broken = self._content().replace("TOTAL PAY SALARY :23954.30",
                                         "TOTAL PAY SALARY :99999.99")
        db = FakeSupabase()
        result = ingest_monthly_email(db, self._email(broken))
        self.assertEqual(result["status"], "error")
        self.assertEqual(db.rows(PAYOUTS_TABLE), [])
        self.assertEqual(db.rows(MONTHLY_REPORT_TABLE), [])

    def test_the_rejection_carries_an_ops_alert_with_both_numbers(self):
        broken = self._content().replace("TOTAL PAY SALARY :23954.30",
                                         "TOTAL PAY SALARY :99999.99")
        result = ingest_monthly_email(FakeSupabase(), self._email(broken))
        alert = result["alert"]
        self.assertIn("salary_block", alert)
        self.assertIn("RM23,954.30", alert)   # parsed
        self.assertIn("RM99,999.99", alert)   # printed
        self.assertIn("NOTHING was stored", alert)

    def test_failures_are_structured_for_the_alert(self):
        broken = self._content().replace("TOTAL PAY SALARY :23954.30",
                                         "TOTAL PAY SALARY :99999.99")
        result = ingest_monthly_email(FakeSupabase(), self._email(broken))
        failure = result["failures"][0]
        self.assertEqual(failure["block"], "salary_block")
        self.assertIn("parsed", failure)
        self.assertIn("printed", failure)
