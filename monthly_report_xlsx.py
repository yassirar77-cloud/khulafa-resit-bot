"""Write the P&L / Cash Reconciliation / Menu Hygiene sheets into the workbook.

Extends ``Khulafa_[OUTLET]_[MONTH]_[YEAR]_Wastage_Report.xlsx`` in place when it
exists, creating it when it does not. Only the three sheets this module owns are
touched — an existing Cost Calculator sheet (fed by ``/wastage_purchases``) and
anything else in the workbook survive untouched, because the owner edits those
by hand and losing that work to a report re-run would be unforgivable.

The three sheets are rewritten wholesale on each run rather than appended to, so
re-running a close never leaves last run's numbers below this run's.

Two presentation rules carry meaning rather than style:

* When ``outlet_config`` is missing, the bottom line is labelled
  ``CONTRIBUTION (pre-rent/utilities/central payroll)`` and the words "net
  profit" appear nowhere. A reader who skims to the last row must not come away
  believing an outlet is profitable when rent has not been subtracted.
* A wastage variance that could not be computed prints its REASON in the
  variance column instead of a number. An empty cell reads as zero; the text
  reads as "we do not know", which is the true state.
"""
from __future__ import annotations

import logging
import re
from typing import Any

logger = logging.getLogger(__name__)

SHEET_PL = "P&L"
SHEET_CASH = "Cash Reconciliation"
SHEET_HYGIENE = "Menu Hygiene"
SHEET_WASTAGE = "Wastage Variance"

OWNED_SHEETS = (SHEET_PL, SHEET_CASH, SHEET_HYGIENE, SHEET_WASTAGE)

_MONTHS = ("JANUARY", "FEBRUARY", "MARCH", "APRIL", "MAY", "JUNE", "JULY",
           "AUGUST", "SEPTEMBER", "OCTOBER", "NOVEMBER", "DECEMBER")

_FILENAME_SAFE_RE = re.compile(r"[^A-Za-z0-9]+")


def report_filename(outlet: Any, period: Any) -> str:
    """``Khulafa_KLANG_B_EMAS_JULY_2026_Wastage_Report.xlsx``."""
    token = _FILENAME_SAFE_RE.sub("_", str(outlet or "UNKNOWN")).strip("_").upper() or "UNKNOWN"
    match = re.fullmatch(r"(\d{4})-(\d{2})", str(period or "").strip())
    if match is None:
        return f"Khulafa_{token}_Wastage_Report.xlsx"
    year, month = int(match.group(1)), int(match.group(2))
    month_name = _MONTHS[month - 1] if 1 <= month <= 12 else str(month)
    return f"Khulafa_{token}_{month_name}_{year}_Wastage_Report.xlsx"


def _money(value: Any):
    """Return a float for numeric cells, or the value unchanged for text."""
    if isinstance(value, bool) or value is None:
        return value
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    return value


def _write_rows(ws, rows: list[list], *, bold_rows: set[int] | None = None,
                money_columns: set[int] | None = None):
    """Write rows starting at A1, bolding the requested 1-based row numbers."""
    from openpyxl.styles import Font

    bold_rows = bold_rows or set()
    money_columns = money_columns or set()
    for r_index, row in enumerate(rows, start=1):
        for c_index, value in enumerate(row, start=1):
            cell = ws.cell(row=r_index, column=c_index)
            cell.value = _money(value) if c_index in money_columns else value
            if c_index in money_columns and isinstance(cell.value, float):
                cell.number_format = "#,##0.00"
            if r_index in bold_rows:
                cell.font = Font(bold=True)
    for column_cells in ws.columns:
        longest = max((len(str(c.value)) for c in column_cells if c.value is not None), default=8)
        ws.column_dimensions[column_cells[0].column_letter].width = min(60, max(12, longest + 2))


def _replace_sheet(wb, title: str):
    """Drop and recreate a sheet so a re-run never stacks two months of rows."""
    if title in wb.sheetnames:
        del wb[title]
    return wb.create_sheet(title)


def build_pl_rows(pl: dict, outlet: str, period: str) -> list[list]:
    from monthly_close import format_pl_lines

    rows: list[list] = [
        [f"P&L — {outlet}, {period}", None],
        [None, None],
        ["Line", "RM"],
    ]
    for label, value in format_pl_lines(pl):
        rows.append([label, value])
    rows += [[None, None], ["Reclassifications applied", None]]
    rows.append([
        "(a) 'TOTAL PAY SALARY' block treated as suppliers, not payroll",
        pl["cogs_breakdown"]["supplier_bank_block"],
    ])
    rows.append([
        "(b) 'PAYOUT PINJAM' treated as a recoverable advance, added back",
        pl["pinjam_addback"],
    ])
    if pl["flags"]:
        rows += [[None, None], ["Flags — resolve before relying on this sheet", None]]
        for flag in pl["flags"]:
            rows.append([flag, None])
    return rows


def build_cash_rows(cash: dict, outlet: str, period: str) -> list[list]:
    components = cash.get("components", {})
    rows: list[list] = [
        [f"Cash reconciliation — {outlet}, {period}", None],
        [None, None],
        ["Line", "RM"],
        ["TOTAL SALES", components.get("total_sales")],
        ["less TOTAL BANK SALES", components.get("total_bank_sales")],
        ["less TOTAL FP SALES", components.get("total_fp_sales")],
        ["CASH SALES", cash.get("cash_sales")],
        ["less TOTAL PAYOUTS", components.get("total_payouts")],
        ["EXPECTED BANK-IN (match this against the bank statement)",
         cash.get("expected_bank_in")],
    ]
    if cash.get("flags"):
        rows += [[None, None], ["Flags", None]]
        for flag in cash["flags"]:
            rows.append([flag, None])
    return rows


def build_hygiene_rows(hygiene: dict, outlet: str, period: str) -> list[list]:
    rows: list[list] = [
        [f"Menu hygiene — {outlet}, {period}", None, None, None],
        [None, None, None, None],
        ["Distinct SKUs", hygiene.get("distinct_skus"), None, None],
        ["Duplicate SKU groups", hygiene.get("duplicate_count"), "RM at stake",
         hygiene.get("duplicate_amount")],
        ["Dead SKUs (qty <= 5)", hygiene.get("dead_count"), "RM",
         hygiene.get("dead_amount")],
        ["Open-price / FOC buttons", hygiene.get("leakage_count"), "RM",
         hygiene.get("leakage_amount")],
        [None, None, None, None],
        ["DUPLICATE SKUs — « » marks leading/trailing or double spaces", None, None, None],
        ["Fold key", "Spelling as printed", "Qty", "RM"],
    ]
    for group in hygiene.get("duplicates", []):
        for index, variant in enumerate(group["variants"]):
            rows.append([
                group["key"] if index == 0 else None,
                variant["display"], variant["qty"], variant["amount"],
            ])
    rows += [[None, None, None, None],
             ["OPEN-PRICE / FOC BUTTONS (leakage surface)", None, None, None],
             ["Item", "Txns", "RM", None]]
    for entry in hygiene.get("leakage", []):
        rows.append([entry["display"], entry["qty"], entry["amount"], None])
    rows += [[None, None, None, None],
             ["DEAD SKUs", None, None, None],
             ["Item", "Qty", "RM", None]]
    for entry in hygiene.get("dead", []):
        rows.append([entry["display"], entry["qty"], entry["amount"], None])
    return rows


def build_wastage_rows(wastage: dict, outlet: str, period: str) -> list[list]:
    rows: list[list] = [
        [f"Wastage variance — {outlet}, {period}", None, None, None, None, None, None],
        [None, None, None, None, None, None, None],
        ["Ingredient", "Theoretical", "Unit", "Purchased", "Variance %",
         "Verdict / reason", "Rule breakdown"],
    ]
    for row in wastage.get("rows", []):
        variance = (
            f"{row['variance_pct']:+.1f}%" if row.get("variance_pct") is not None
            else (row.get("reason") or "—")
        )
        breakdown = ", ".join(
            f"{k} {v:g}" for k, v in sorted((row.get("ingredients") or {}).items())
        )
        rows.append([
            row["canonical_item"], row.get("theoretical_qty"),
            row.get("theoretical_unit") or row.get("purchased_unit"),
            row.get("purchased_qty"), variance,
            row.get("verdict") if row.get("variance_pct") is not None
            else f"{row.get('verdict')} — {row.get('reason') or ''}",
            breakdown,
        ])
    if wastage.get("flagged_rows"):
        rows += [
            [None, None, None, None, None, None, None],
            [f"{wastage['flagged_rows']} purchase line(s) worth "
             f"RM{wastage['flagged_cost']:.2f} could not be converted to a base "
             f"quantity — those ingredients show no %",
             None, None, None, None, None, None],
        ]
    return rows


def write_monthly_sheets(
    path: str, *, outlet: str, period: str,
    pl: dict, cash: dict, hygiene: dict, wastage: dict | None = None,
) -> str:
    """Write the sheets into ``path``, preserving any other sheets present.

    Returns the path written. Raises ``ImportError`` with an actionable message
    when openpyxl is unavailable rather than failing deep inside the library.
    """
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError as exc:  # pragma: no cover - dependency guard
        raise ImportError(
            "openpyxl is required to write the monthly workbook — "
            "pip install openpyxl (it is in requirements.txt)"
        ) from exc

    import os

    if os.path.exists(path):
        wb = load_workbook(path)
        logger.info("monthly xlsx: extending %s (existing sheets: %s)", path, wb.sheetnames)
    else:
        wb = Workbook()
        # A fresh Workbook starts with an empty default sheet; drop it only if
        # our sheets will replace it, never a sheet the owner created.
        default = wb.active
        if default is not None and default.max_row == 1 and default.max_column == 1 \
                and default["A1"].value is None:
            del wb[default.title]
        logger.info("monthly xlsx: creating %s", path)

    _write_rows(_replace_sheet(wb, SHEET_PL), build_pl_rows(pl, outlet, period),
                bold_rows={1, 3}, money_columns={2})
    _write_rows(_replace_sheet(wb, SHEET_CASH), build_cash_rows(cash, outlet, period),
                bold_rows={1, 3, 9}, money_columns={2})
    _write_rows(_replace_sheet(wb, SHEET_HYGIENE), build_hygiene_rows(hygiene, outlet, period),
                bold_rows={1, 9}, money_columns={4})
    if wastage is not None:
        _write_rows(_replace_sheet(wb, SHEET_WASTAGE),
                    build_wastage_rows(wastage, outlet, period),
                    bold_rows={1, 3}, money_columns=set())

    wb.save(path)
    logger.info("monthly xlsx: wrote %s -> %s", OWNED_SHEETS, path)
    return path
